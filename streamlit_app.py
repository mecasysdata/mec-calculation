import streamlit as st
import pandas as pd
import numpy as np
import datetime
import math
import re
import joblib

# -----------------------------------------
# 1. CONFIG & SESSION STATE
# -----------------------------------------
st.set_page_config(layout="wide", page_title="MEC Calculation")

if "kosik" not in st.session_state:
    st.session_state.kosik = []

if "stary_item" not in st.session_state:
    st.session_state.stary_item = ""

if "aktualny_pocet_kusov" not in st.session_state:
    st.session_state.aktualny_pocet_kusov = 1

if "cas_potvrdeny" not in st.session_state:
    st.session_state.cas_potvrdeny = False
if "schvaleny_cas" not in st.session_state:
    st.session_state.schvaleny_cas = 3.0

if "cena_potvrdena" not in st.session_state:
    st.session_state.cena_potvrdena = False
if "schvalena_cena" not in st.session_state:
    st.session_state.schvalena_cena = 3.0

# -----------------------------------------
# 2. LOAD DATA FROM GOOGLE SHEETS
# -----------------------------------------
@st.cache_data(ttl=600)
def load_data_from_url(url: str) -> pd.DataFrame:
    try:
        df = pd.read_csv(url)
        df.columns = df.columns.str.strip().str.lower()
        return df
    except Exception as e:
        st.error(f"Chyba pri načítavaní dát: {e}")
        return pd.DataFrame()

SHEET_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSuHQWbpryWNerWr8aKKheHbzTPhXI6lS7YH1sL5zwFIIzLfpTZz47acYua2efVqEcfxMBe5wnjue/pub?gid=0&single=true&output=csv"
MATERIAL_SHEET_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQf4EiqZt1grkazJgfYWVhG0M8FGLNCjaGk6dcXhO3r04JQuZ9Qxv1jelDo3c8hBLy7Ny5C1pZqvbfS/pub?output=csv"
SHEET_KOOP_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRfPBZ4TCpQyiqybU0ADu3AMwHCi2qOKifQAOnnTWnorVNJ1SVxtN6zJzXthOxCVwtXWp__Bp_-nto0/pub?gid=1180392224&single=true&output=csv"

df = load_data_from_url(SHEET_URL)
df_mat = load_data_from_url(MATERIAL_SHEET_URL)
df_koop = load_data_from_url(SHEET_KOOP_URL)

if df.empty or df_mat.empty or df_koop.empty:
    st.error("❌ Kritické dáta sa nepodarilo načítať. Skontroluj Google Sheets.")
    st.stop()

# -----------------------------------------
# 3. LOAD STV MODEL (KR zatiaľ vypnutý)
# -----------------------------------------
@st.cache_resource
def load_stv_model():
    try:
        model = joblib.load("model_stv.pkl")
        encoder = joblib.load("encoder_stv.pkl")
        return model, encoder
    except Exception as e:
        st.warning(f"⚠️ STV model sa nepodarilo načítať: {e}")
        return None, None

model_stv, encoder_stv = load_stv_model()

# -----------------------------------------
# 4. STV SAFETY PIPELINE (log1p, whitelist, type casting)
# -----------------------------------------
POVOLENE_SUBCATS = [
    "UNALL", "LOWAL", "ALLOYED", "TOOL", "HSS",
    "AUST", "DUPX", "MART", "FERR",
    "CU", "BRASS", "BRONZE", "ALU", "TI", "NI-SPEC",
    "PEEK", "PET-G", "PMMA", "PC", "PUR", "RUBBER",
    "PVC", "POM", "PET", "PA", "PP", "PE",
    "CAST-GG", "CAST-GGG", "CAST-TEMP",
    "OTHER", "OSTATNÉ"
]

def prepare_stv_input(v_narocnost, pocet_kusov, hmotnost_kg, subcategory, plocha_m2, geo_koef):
    log_pocet = np.log1p(max(int(pocet_kusov), 1))

    sub_clean = (subcategory or "").strip().upper()
    if sub_clean not in POVOLENE_SUBCATS:
        sub_clean = "OTHER"

    df_in = pd.DataFrame([{
        "v_narocnost": float(v_narocnost),
        "log_pocet_kusov": float(log_pocet),
        "hmotnost_kg": float(hmotnost_kg),
        "SUBCATEGORY_clean": sub_clean,
        "plocha_m2": float(plocha_m2),
        "geometricky_koeficient": float(geo_koef),
    }])

    return df_in

def predict_stv_time(model, encoder, v_narocnost, pocet_kusov, hmotnost_kg, subcategory, plocha_m2, geo_koef):
    df_in = prepare_stv_input(
        v_narocnost, pocet_kusov, hmotnost_kg, subcategory, plocha_m2, geo_koef
    )

    if encoder is not None:
        df_in[["SUBCATEGORY_clean"]] = encoder.transform(df_in[["SUBCATEGORY_clean"]])

    return float(model.predict(df_in)[0])
# -----------------------------------------
# 5. ZÁKAZNÍK
# -----------------------------------------
col1, col2, col3, col4 = st.columns(4)

with col1:
    datum = st.date_input("Dátum", datetime.date.today())

with col2:
    ponuka = st.text_input("Označenie CP")

zoznam_zakaznikov = sorted(df['zakaznik'].dropna().unique()) if 'zakaznik' in df.columns else []
moznosti_zakaznikov = ["+ Pridať nového zákazníka"] + zoznam_zakaznikov

with col3:
    vyber = st.selectbox("Názov Zákazníka", moznosti_zakaznikov)

zakaznik = ""
krajina_hodnota = ""
lojalita = 0.5

if vyber == "+ Pridať nového zákazníka":
    with col3:
        zakaznik = st.text_input("Meno nového zákazníka", key="new_cust_name")
    with col4:
        krajina_hodnota = st.text_input("Krajina Zákazníka (manuálne)", key="new_cust_country")
else:
    filter_zak = df[df['zakaznik'] == vyber]
    if not filter_zak.empty:
        data_zakaznika = filter_zak.iloc[0]
        zakaznik = vyber
        krajina_hodnota = str(data_zakaznika.get('krajina', 'Neznáma'))
        lojalita = float(data_zakaznika.get('lojalita', 0.5))
    else:
        st.warning(f"Zákazník {vyber} nemá v tabuľke priradené dáta.")

    with col4:
        st.text_input("Krajina Zákazníka", value=krajina_hodnota, disabled=True, key="disabled_country")

st.divider()

# -----------------------------------------
# 6. ITEM & GEOMETRIA
# -----------------------------------------
col5, col6, col7, col8, col9, col10, col11 = st.columns(7)

with col5:
    item = st.text_input("ITEM", key="item_input")

# Reset logiky pri zmene ITEMu
if item != st.session_state.stary_item:
    if st.session_state.stary_item != "" and item.strip() != "":
        st.session_state.aktualny_pocet_kusov = 1

        kluce_na_vymazanie = [
            "pocet_input", "narocnost_input", "tvar_input",
            "d_kr", "l_kr", "d_stv", "s_stv", "v_stv",
            "mat_select", "akost_multi", "man_akost_chk",
            "polo_inteligent", "koop_main_checkbox",
            "manual_rho", "mat_k", "druh_k",
            "vystupny_cas_input", "vystupna_cena_input"
        ]

        for k in kluce_na_vymazanie:
            if k in st.session_state:
                del st.session_state[k]

        st.session_state.cas_potvrdeny = False
        st.session_state.schvaleny_cas = 3.0
        st.session_state.cena_potvrdena = False
        st.session_state.schvalena_cena = 3.0

        st.session_state.stary_item = item
        st.rerun()
    else:
        st.session_state.stary_item = item

with col6:
    predvoleny_pocet = st.session_state.get("aktualny_pocet_kusov", 1)
    pocet_kusov = st.number_input(
        "Počet kusov",
        min_value=1,
        value=int(predvoleny_pocet),
        key="pocet_input"
    )
    st.session_state.aktualny_pocet_kusov = pocet_kusov

with col7:
    narocnost = st.selectbox("Náročnosť", options=[1, 2, 3, 4, 5], key="narocnost_input")

with col8:
    tvar_item = st.selectbox("Tvar položky", options=["STV", "KR"], key="tvar_input")

# Geometrické vstupy
d = l = s = v = 0.0

if tvar_item == "KR":
    with col9:
        d = st.number_input("D(mm)", min_value=0.0, format="%.1f", key="d_kr")
    with col10:
        l = st.number_input("L(mm)", min_value=0.0, format="%.1f", key="l_kr")
else:
    with col9:
        d = st.number_input("D/P(mm)", min_value=0.0, format="%.1f", key="d_stv")
    with col10:
        s = st.number_input("S(mm)", min_value=0.0, format="%.1f", key="s_stv")
    with col11:
        v = st.number_input("V(mm)", min_value=0.0, format="%.1f", key="v_stv")

st.divider()

# -----------------------------------------
# 7. MATERIÁL & POLOTOVAR
# -----------------------------------------
def get_sorted_dims(a, b, c):
    try:
        return sorted([float(a), float(b), float(c)], reverse=True)
    except:
        return [0.0, 0.0, 0.0]

col_m1, col_m2, col_m3 = st.columns([2, 3, 3])

with col_m1:
    zoznam_materialov = sorted(df_mat['material'].dropna().unique()) if 'material' in df_mat else []
    material_vyber = st.selectbox("Materiál", zoznam_materialov, key="mat_select")

with col_m2:
    df_f_akost = df_mat[df_mat['material'] == material_vyber]
    filtr_akosti_vsetky = sorted(df_f_akost['akost'].dropna().astype(str).unique()) if 'akost' in df_f_akost else []
    akost_vyber_list = st.multiselect("Výber akostí", options=filtr_akosti_vsetky, key="akost_multi")
    manual_akost_check = st.checkbox("+ Iná akosť (manuálne)", key="man_akost_chk")

vhodne_moznosti = []

if manual_akost_check:
    zoznam_na_vyber = ["+ Pridať nový/iný polotovar"]

else:
    df_relevant = df_mat[
        (df_mat['material'] == material_vyber) &
        (df_mat['akost'].astype(str).isin(akost_vyber_list))
    ].copy()

    if tvar_item == "KR" and 'názov' in df_relevant:
        df_relevant = df_relevant[df_relevant['názov'].str.contains("KR|6HR|TR", case=False, na=False)]

    if not df_relevant.empty:
        df_relevant["sort_key"] = df_relevant.apply(
            lambda r: get_sorted_dims(r['rozmer1'], r['rozmer2'], r['rozmer3']),
            axis=1
        )
        df_relevant = df_relevant.sort_values(by="sort_key")

        for _, r in df_relevant.iterrows():
            label = (
                f"[{r.get('akost','')}] {r.get('názov','')} | "
                f"{r.get('rozmer1',0)}x{r.get('rozmer2',0)}x{r.get('rozmer3',0)} | "
                f"Cena: {r.get('cena',0)} €/bm"
            )
            vhodne_moznosti.append({
                "label": label,
                "cena": float(r.get("cena", 0)),
                "akost_povodna": str(r.get("akost", ""))
            })

    zoznam_na_vyber = [x["label"] for x in vhodne_moznosti] + ["+ Pridať nový/iný polotovar"]

with col_m3:
    idx_start = len(zoznam_na_vyber) - 1 if (manual_akost_check or not vhodne_moznosti) else 0
    vybrany_polo_str = st.selectbox("Výber polotovaru", zoznam_na_vyber, index=idx_start, key="polo_inteligent")

cena_polotovaru = 0.0
relevantna_akost = ""

if vybrany_polo_str == "+ Pridať nový/iný polotovar":
    c1, c2, c3 = st.columns(3)

    with c1:
        povodna = akost_vyber_list[0] if akost_vyber_list else ""
        nova_akost = st.text_input("Názov akosti", value=povodna, key="input_nov_akost")

    with c2:
        cena_polotovaru = st.number_input("Cena (€/bm)", min_value=0.0, format="%.2f", key="input_nov_cena")

    with c3:
        nazov_pol = st.text_input("Názov polotovaru", value="MANUAL", key="input_nov_nazov")

    relevantna_akost = nova_akost.upper().replace(" ", "").strip()

else:
    vybrany_objekt = next((x for x in vhodne_moznosti if x["label"] == vybrany_polo_str), None)
    if vybrany_objekt:
        cena_polotovaru = vybrany_objekt["cena"]
        relevantna_akost = vybrany_objekt["akost_povodna"].upper().replace(" ", "").strip()

# Výpočet ceny materiálu
dlzka_pre_vypocet = l if tvar_item == "KR" else d
cena_mat_kus = (dlzka_pre_vypocet / 1000) * cena_polotovaru
# -----------------------------------------
# 8. GEOMETRIA & HMOTNOSŤ
# -----------------------------------------

if tvar_item == "KR":
    # Pôvodná KR geometria (KR model zatiaľ nemáš)
    plocha_prierezu = (math.pi * (d ** 2)) / 4
    povrch_celkovy_mm2 = (2 * plocha_prierezu) + (math.pi * d * l)
    hmotnost_kusu = (plocha_prierezu * l * hustota) / 1e9

else:
    # -----------------------------------------
    # STV GEOMETRIA – TVOJA PRESNÁ Z TRÉNINGU
    # -----------------------------------------
    dlzka = d
    sirka = s
    vyska = v

    plocha_m2 = (dlzka / 1000) * (sirka / 1000)
    objem_m3 = plocha_m2 * (vyska / 1000)
    hmotnost_kusu = objem_m3 * hustota

    geo_koef = plocha_m2 / (hmotnost_kusu + 0.001)

    # Pre info panel (aby sa nič nerozbilo)
    plocha_prierezu = plocha_m2 * 1_000_000
    povrch_celkovy_mm2 = plocha_prierezu

plocha_prierez_dm2 = povrch_celkovy_mm2 / 10000
hmotnost_celkom = hmotnost_kusu * pocet_kusov

# -----------------------------------------
# 9. KOOPERÁCIA & NÁKLADY
# -----------------------------------------
st.write("---")
rk1, rk2, rk3, rk4, rk5, rk6, rk7 = st.columns([0.8, 1.5, 1.5, 1.2, 1.2, 1.2, 1.5])

with rk1:
    je_kooperacia = st.checkbox("Koop?", value=False, key="koop_main_checkbox")

cena_kooperacia = 0.0

if je_kooperacia:
    zoznam_vsetkych_mat_koop = sorted(df_koop['material'].dropna().unique()) if 'material' in df_koop else []
    try:
        default_idx = zoznam_vsetkych_mat_koop.index(material_vyber)
    except:
        default_idx = 0

    with rk3:
        vybrany_mat_koop = st.selectbox("Mat. koop.", zoznam_vsetkych_mat_koop, index=default_idx, key="mat_k")

    with rk2:
        df_f_koop = df_koop[df_koop['material'] == vybrany_mat_koop]
        mozne_operacie = sorted(df_f_koop['druh'].dropna().unique()) if 'druh' in df_f_koop else []
        vybrany_druh = st.selectbox("Druh koop.", mozne_operacie, key="druh_k")

    filter_koop_row = df_koop[
        (df_koop['druh'] == vybrany_druh) &
        (df_koop['material'] == vybrany_mat_koop)
    ]

    if not filter_koop_row.empty:
        riadok_koop = filter_koop_row.iloc[0]
        tarifa = float(riadok_koop.get('tarifa', 0))
        jednotka = str(riadok_koop.get('jednotka', '')).strip().lower()
        min_obj = float(riadok_koop.get('minimum', 0))

        if jednotka == "kg":
            zaklad = hmotnost_kusu
        elif jednotka == "dm2":
            zaklad = plocha_prierez_dm2
        else:
            zaklad = 1

        vyp_cena = tarifa * zaklad
        cena_kooperacia = max(vyp_cena, min_obj / pocet_kusov)
    else:
        st.warning("Pre zvolenú kombináciu druhu a materiálu kooperácie neboli nájdené ceny.")

else:
    with rk2:
        st.text_input("Druh koop.", "-", disabled=True)
    with rk3:
        st.text_input("Mat. koop.", "-", disabled=True)

vstupne_naklady = cena_mat_kus + cena_kooperacia

with rk4:
    st.metric("Cena/bm", f"{cena_polotovaru:.2f} €")
with rk5:
    st.metric("Mat./kus", f"{cena_mat_kus:.3f} €")
with rk6:
    st.metric("Koop./kus", f"{cena_kooperacia:.3f} €")
with rk7:
    st.metric("VSTUPNÉ NÁKLADY", f"{vstupne_naklady:.3f} €", delta=f"{hmotnost_kusu:.2f} kg", delta_color="off")

# -----------------------------------------
# INFO PANEL
# -----------------------------------------
st.markdown(
    f"""
    <div style="background-color:#f1f3f6;padding:10px;border-radius:5px;font-size:0.85em;color:#555;">
    <strong>Použitá akosť:</strong> {relevantna_akost} |
    <strong>Subcategory:</strong> {subcategory} |
    <strong>Hustota:</strong> {hustota:.0f} kg/m³ |
    <strong>Plocha prierezu:</strong> {plocha_prierezu:.2f} mm² |
    <strong>Hmotnosť:</strong> {hmotnost_kusu:.3f} kg |
    <strong>Povrch:</strong> {plocha_prierez_dm2:.3f} dm²
    </div>
    """,
    unsafe_allow_html=True
)

# -----------------------------------------
# 10. MEC AI – STV MODEL
# -----------------------------------------
st.divider()
st.subheader("🤖 MEC AI")

model_predikcia_cas = 3.0
model_predikcia_cena = 3.0

if tvar_item == "STV" and model_stv is not None:
    try:
        # Použijeme tvoju presnú geometriu
        cas_pred = predict_stv_time(
            model_stv,
            encoder_stv,
            v_narocnost=narocnost,
            pocet_kusov=pocet_kusov,
            hmotnost_kg=hmotnost_kusu,
            subcategory=subcategory,
            plocha_m2=plocha_m2,
            geo_koef=geo_koef
        )
        model_predikcia_cas = cas_pred

    except Exception as e:
        st.warning(f"Predikcia STV modelu zlyhala: {e}")
        model_predikcia_cas = 3.0

# KR model zatiaľ nemáš → fallback
if tvar_item == "KR":
    model_predikcia_cas = 3.0

# -----------------------------------------
# AI PANEL – UI
# -----------------------------------------
ai_col1, ai_col2, ai_col3, ai_col4, ai_col5 = st.columns(5)

with ai_col1:
    vystupny_cas = st.number_input(
        "Výrobný čas /ks [min]",
        min_value=0.0,
        value=float(model_predikcia_cas),
        format="%.2f",
        key="vystupny_cas_input"
    )

with ai_col2:
    st.write("")
    if st.button("✅ Schváliť výrobný čas", type="secondary", use_container_width=True):
        st.session_state.schvaleny_cas = vystupny_cas
        st.session_state.cas_potvrdeny = True
        st.rerun()

with ai_col3:
    if st.session_state.cas_potvrdeny:
        vystupna_cena = st.number_input(
            "Predikovaná cena /ks (€)",
            min_value=0.0,
            value=float(model_predikcia_cena),
            format="%.2f",
            key="vystupna_cena_input"
        )
    else:
        st.info("💡 Čaká sa na schválenie času...")
        vystupna_cena = 0.0

with ai_col4:
    st.write("")
    if st.button("✅ Schváliť cenu", type="secondary", use_container_width=True, disabled=not st.session_state.cas_potvrdeny):
        st.session_state.schvalena_cena = vystupna_cena
        st.session_state.cena_potvrdena = True
        st.rerun()

with ai_col5:
    st.write("")
    if st.button("🛒 Pridať item do košíka", type="primary", use_container_width=True, disabled=not st.session_state.cena_potvrdena):
        if item.strip() == "":
            st.warning("Zadaj názov ITEMu pred pridaním.")
        else:
            if tvar_item == "KR":
                rozmery_formatted = f"{d:.1f} x {l:.1f}"
            else:
                rozmery_formatted = f"{d:.1f} x {s:.1f} x {v:.1f}"

            nova_polozka = {
                "ITEM": item,
                "Počet kusov": pocet_kusov,
                "Materiál": material_vyber,
                "Akosť": relevantna_akost,
                "Rozmery": rozmery_formatted,
                "Popis polotovaru": vybrany_polo_str,
                "Výrobný čas (min/ks)": round(st.session_state.schvaleny_cas, 2),
                "Model Cena (€/ks)": round(st.session_state.schvalena_cena, 2),
                "Mat. / kus (€)": round(cena_mat_kus, 3),
                "Koop. / kus (€)": round(cena_kooperacia, 3),
                "Vstupné náklady (€/ks)": round(vstupne_naklady, 3),
                "Celkom za položku (€)": round(vstupne_naklady * pocet_kusov, 2)
            }

            st.session_state.kosik.append(nova_polozka)
            st.success(f"Položka '{item}' pridaná!")
            st.rerun()
# -----------------------------------------
# 12. KOŠÍK, PDF, EXCEL, SHEET EXPORT
# -----------------------------------------

if st.session_state.kosik:
    st.write("---")
    st.subheader(f"📋 Aktuálny zoznam položiek v ponuke (Počet: {len(st.session_state.kosik)})")

    df_kosik = pd.DataFrame(st.session_state.kosik)
    st.dataframe(df_kosik, use_container_width=True, hide_index=True)

    celkova_suma = df_kosik["Celkom za položku (€)"].sum()

    col_sum1, col_sum2 = st.columns([5, 3])
    with col_sum2:
        st.metric("CELKOVÁ CENA PONUKY", f"{celkova_suma:.2f} €")

col_pdf, col_save, col_reset = st.columns(3)

# -----------------------------------------
# PDF EXPORT
# -----------------------------------------
with col_pdf:
    if not st.session_state.kosik:
        st.info("🛒 Pre stiahnutie PDF pridajte položky do košíka.")
    else:
        try:
            import unicodedata
            from fpdf import FPDF
            import io
            import openpyxl

            def odstran_diakritiku(text):
                if not text:
                    return ""
                nfkd = unicodedata.normalize("NFKD", str(text))
                return "".join([c for c in nfkd if not unicodedata.combining(c)])

            class CP_PDF(FPDF):
                def header(self):
                    self.set_font("Helvetica", "B", 12)
                    self.cell(0, 10, "CENOVÁ PONUKA (MEC Calculation)", ln=True, align="C")
                    self.ln(3)

                def footer(self):
                    self.set_y(-15)
                    self.set_font("Helvetica", "I", 8)
                    self.cell(0, 10, f"Strana {self.page_no()}", align="C")

            with st.spinner("⏳ Generujem PDF..."):
                pdf = CP_PDF()
                pdf.add_page()
                pdf.set_font("Helvetica", size=9)

                c_ponuka = odstran_diakritiku(ponuka)
                c_zakaznik = odstran_diakritiku(zakaznik)
                c_krajina = odstran_diakritiku(krajina_hodnota)
                c_datum = datum.strftime("%d.%m.%Y")

                pdf.set_font("Helvetica", "B", 10)
                pdf.cell(0, 6, f"Označenie CP: {c_ponuka}", ln=True)
                pdf.cell(0, 6, f"Dátum: {c_datum}", ln=True)
                pdf.cell(0, 6, f"Zákazník: {c_zakaznik} ({c_krajina})", ln=True)
                pdf.ln(4)

                # Hlavička tabuľky
                pdf.set_font("Helvetica", "B", 8)
                headers = [
                    ("ITEM", 25), ("Materiál", 15), ("Akosť", 12),
                    ("Polotovar", 30), ("Rozmery", 30), ("Ks", 12),
                    ("Mat./ks", 18), ("Koop./ks", 18), ("Vstup", 18),
                    ("Cena/ks", 20), ("Spolu", 25)
                ]
                for h, w in headers:
                    pdf.cell(w, 7, h, border=1, align="C")
                pdf.ln()

                pdf.set_font("Helvetica", size=8)
                for p in st.session_state.kosik:
                    row = [
                        p["ITEM"], p["Materiál"], p["Akosť"],
                        p["Popis polotovaru"], p["Rozmery"],
                        str(p["Počet kusov"]),
                        f"{p['Mat. / kus (€)']:.3f}",
                        f"{p['Koop. / kus (€)']:.3f}",
                        f"{p['Vstupné náklady (€/ks)']:.3f}",
                        f"{p['Model Cena (€/ks)']:.2f}",
                        f"{p['Celkom za položku (€)']:.2f}",
                    ]
                    for (h, w), val in zip(headers, row):
                        pdf.cell(w, 7, odstran_diakritiku(val), border=1)
                    pdf.ln()

                pdf.ln(4)
                pdf.set_font("Helvetica", "B", 11)
                pdf.cell(0, 10, f"CELKOVÁ CENA: {celkova_suma:.2f} €", ln=True, align="R")

                pdf_data = pdf.output(dest="S").encode("latin1")

            st.download_button(
                label="📄 1. Stiahnuť PDF ponuku",
                data=pdf_data,
                file_name=f"Cenova_ponuka_{c_ponuka or 'MEC'}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

            # -----------------------------------------
            # EXCEL EXPORT
            # -----------------------------------------
            output = io.BytesIO()
            df_export = pd.DataFrame(st.session_state.kosik)

            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_export.to_excel(writer, index=False, sheet_name="Ponuka")
                writer.save()

            excel_data = output.getvalue()

            st.download_button(
                label="📊 Stiahnuť Excel ponuku",
                data=excel_data,
                file_name=f"Cenova_ponuka_{c_ponuka or 'MEC'}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        except Exception as e:
            st.error(f"Chyba pri generovaní PDF/Excel: {e}")

# -----------------------------------------
# GOOGLE SHEET EXPORT
# -----------------------------------------
with col_save:
    if not st.session_state.kosik:
        st.info("💾 Pre uloženie naplňte košík.")
    else:
        if st.button("💾 2. Uložiť a uzatvoriť ponuku", type="primary", use_container_width=True):

            if not ponuka.strip():
                st.error("❌ Zadaj 'Označenie CP'!")
            elif not zakaznik.strip():
                st.error("❌ Zadaj alebo vyber zákazníka!")
            else:
                import requests

                riadky = []
                for p in st.session_state.kosik:
                    riadky.append({
                        "Dátum CP": datum.strftime("%d.%m.%Y"),
                        "Číslo CP": ponuka,
                        "Zákazník": zakaznik,
                        "Krajina": krajina_hodnota,
                        "Lojalita": lojalita,
                        "ITEM": p["ITEM"],
                        "Tvar": "KR" if " x " in p["Rozmery"] and p["Rozmery"].count("x") == 1 else "STV",
                        "Materiál": p["Materiál"],
                        "Akosť": p["Akosť"],
                        "Rozmery": p["Rozmery"],
                        "Výrobný čas (min/ks)": p["Výrobný čas (min/ks)"],
                        "Jednotková cena (€/ks)": p["Model Cena (€/ks)"],
                        "Počet kusov": p["Počet kusov"],
                        "Cena položky spolu (€)": p["Celkom za položku (€)"],
                    })

                URL_SCRIPTU = "https://script.google.com/macros/s/AKfycbwx7sAeUheQf1dm2r6k7jTslD9ufhq2yk1OWZXWjxVkeZOttVI949GIiPGx8l1B3cIP/exec"

                with st.spinner("⏳ Zapisujem do Google Sheet..."):
                    try:
                        odpoved = requests.post(URL_SCRIPTU, json=riadky)
                        if "success" in odpoved.text.lower():
                            st.success(f"🎉 Ponuka '{ponuka}' bola uložená!")
                        else:
                            st.error(f"Chyba skriptu: {odpoved.text}")
                    except Exception as e:
                        st.error(f"Chyba pripojenia: {e}")

# -----------------------------------------
# RESET APLIKÁCIE
# -----------------------------------------
with col_reset:
    if st.button("🆕 Nová cenová ponuka", type="secondary", use_container_width=True):
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.session_state.kosik = []
        st.session_state.stary_item = ""
        st.session_state.aktualny_pocet_kusov = 1
        st.session_state.cas_potvrdeny = False
        st.session_state.cena_potvrdena = False
        st.rerun()
